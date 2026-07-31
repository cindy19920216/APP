# -*- coding: utf-8 -*-
"""
KOSPI 대형주 몇 종목(외신 커버리지가 실제로 있는 종목만)의 뉴스를 매일 자동 수집한다.

Bing News RSS로 종목당 최대 5건을 가져오되, 영어(외신, mkt=en-US) 검색을 우선하고
5건이 안 채워지면 한국어(국내 매체, mkt=ko-KR) 검색으로 부족분을 채운다.

Google News RSS도 처음에 검토했으나, 그쪽 링크(`news.google.com/rss/articles/...`)는
requests로는 실제 기사 페이지로 리다이렉트되지 않고(JS 리다이렉트라 항상 구글
자체 안내 페이지만 나옴) 그 결과 모든 기사의 요약이 "Comprehensive up-to-date news
coverage..." 같은 구글의 고정 문구로 동일하게 나오는 문제가 실측으로 확인됐다.
Bing News RSS는 `<description>`에 실제 기사 스니펫을 직접 담아주고, 링크도
`apiclick.aspx?...&url={실제기사URL}` 형태라 그 자리에서 바로 실제 URL을 추출할 수
있어 이 방식으로 전환했다 — 기사 페이지를 다시 열어볼 필요가 없다.

Gemini로 한글 제목/요약을 만들어 news/종목뉴스_YYYYMMDD.xlsx에 저장한다. 빅테크뉴스
(scheduled-tasks/bigtech-news-daily) 엑셀 스키마에 "티커" 컬럼만 추가한 형태 —
src/app/api/stock-news/route.ts가 이 파일을 읽어 종목 상세 뉴스 탭에 서빙한다.

GEMINI_API_KEY가 없거나 번역 호출이 실패해도 수집 자체는 계속 진행하고 한글 필드만
원문으로 채운다 (기존 fetch_charts.py와 동일한 원칙).
"""
import json
import os
import re
import sys
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta, timezone
from email.utils import parsedate_to_datetime
from urllib.parse import parse_qs, quote, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
NEWS_DIR = os.path.join(BASE_DIR, "news")
HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"}
TARGET_COUNT = 5
KST = timezone(timedelta(hours=9))

# 외신 커버리지가 실제로 있는 대형주만 대상 — src/app/api/stock-news/route.ts의
# FEATURED_CODES와 반드시 동기화할 것.
FEATURED_STOCKS = [
    {"code": "005930", "name": "삼성전자", "enName": "Samsung Electronics"},
    {"code": "000660", "name": "SK하이닉스", "enName": "SK Hynix"},
    {"code": "005380", "name": "현대차", "enName": "Hyundai Motor"},
    {"code": "373220", "name": "LG에너지솔루션", "enName": "LG Energy Solution"},
    {"code": "207940", "name": "삼성바이오로직스", "enName": "Samsung Biologics"},
]

GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY")
GEMINI_MODEL = "gemini-3.1-flash-lite"
GEMINI_URL = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent"


def clean_html(html: str) -> str:
    if not html:
        return ""
    return BeautifulSoup(html, "html.parser").get_text(strip=True)


def format_pub_date(pub_date_str: str) -> str:
    try:
        return parsedate_to_datetime(pub_date_str).strftime("%Y-%m-%d")
    except Exception:
        return ""


def local_tag(elem) -> str:
    """Bing RSS의 News: 네임스페이스 URI가 요청 쿼리를 포함해 매 요청마다 달라지므로,
    네임스페이스를 무시하고 태그의 로컬 이름(마지막 '}' 뒤)만으로 비교한다."""
    return elem.tag.rsplit("}", 1)[-1]


def child_text(item, name: str) -> str:
    for child in item:
        if local_tag(child) == name:
            return (child.text or "").strip()
    return ""


def extract_real_url(bing_link: str) -> str:
    """Bing의 apiclick.aspx 추적 링크에서 실제 기사 URL(url= 쿼리파라미터)을 꺼낸다.
    형식이 안 맞으면(추적 링크가 아닌 경우 등) 원래 링크를 그대로 반환."""
    try:
        qs = parse_qs(urlparse(bing_link).query)
        real = qs.get("url", [None])[0]
        return real or bing_link
    except Exception:
        return bing_link


def fetch_rss_items(query: str, mkt: str, limit: int = 8) -> list:
    url = f"https://www.bing.com/news/search?q={quote(query)}&format=RSS&mkt={mkt}"
    try:
        resp = requests.get(url, headers=HEADERS, timeout=20)
        resp.raise_for_status()
        root = ET.fromstring(resp.text)
    except (requests.RequestException, ET.ParseError) as e:
        print(f"    RSS 조회 실패 ({query}, {mkt}): {e}")
        return []

    items = []
    for item in root.findall(".//item")[:limit]:
        title = child_text(item, "title")
        link = extract_real_url(child_text(item, "link"))
        if not title or not link:
            continue
        items.append({
            "title": title,
            "link": link,
            "date": format_pub_date(child_text(item, "pubDate")),
            "media": child_text(item, "Source"),
            "desc": clean_html(child_text(item, "description")),
        })
    return items


def collect_for_stock(stock: dict) -> list:
    seen_links = set()
    selected = []

    foreign = fetch_rss_items(f"{stock['enName']} stock", "en-US")
    for it in foreign:
        if it["link"] in seen_links:
            continue
        seen_links.add(it["link"])
        selected.append(it)
        if len(selected) >= TARGET_COUNT:
            break

    if len(selected) < TARGET_COUNT:
        domestic = fetch_rss_items(f"{stock['name']} 주가", "ko-KR")
        for it in domestic:
            if it["link"] in seen_links:
                continue
            seen_links.add(it["link"])
            selected.append(it)
            if len(selected) >= TARGET_COUNT:
                break

    return selected


def translate_batch(items: list) -> list:
    """items의 title/desc를 한국어로 일괄 번역. 실패/미설정 시 전부 (None, None)."""
    fallback = [(None, None)] * len(items)
    if not GEMINI_API_KEY:
        print("  GEMINI_API_KEY 미설정 - 번역 건너뜀 (원문 그대로 표시됨)")
        return fallback
    if not items:
        return fallback

    payload = [{"title": it["title"], "desc": it["desc"]} for it in items]
    prompt = (
        "다음은 주식 관련 뉴스 기사의 제목과 요약을 담은 JSON 배열이다. "
        "각 항목의 title과 desc를 자연스러운 한국어 증권 뉴스체로 번역하라. "
        "이미 한국어인 항목은 자연스럽게 다듬기만 하라. "
        "고유명사/종목명은 통용되는 한국어 표기를 쓰고, 원문에 없는 내용을 임의로 "
        "덧붙이지 말 것. 설명 외 다른 텍스트 없이, 입력과 같은 개수/순서의 JSON 배열만 "
        '출력하라. 각 원소 형식: {"titleKo": "...", "descKo": "..."}\n\n'
        f"입력:\n{json.dumps(payload, ensure_ascii=False)}"
    )

    try:
        resp = requests.post(
            GEMINI_URL,
            params={"key": GEMINI_API_KEY},
            json={"contents": [{"parts": [{"text": prompt}]}]},
            timeout=60,
        )
        resp.raise_for_status()
        data = resp.json()
        text = data["candidates"][0]["content"]["parts"][0]["text"].strip()
        text = re.sub(r"^```(?:json)?|```$", "", text.strip(), flags=re.MULTILINE).strip()
        translated = json.loads(text)
        if not isinstance(translated, list) or len(translated) != len(items):
            raise ValueError(f"번역 결과 개수 불일치: {len(translated) if isinstance(translated, list) else 'N/A'} != {len(items)}")
        return [(t.get("titleKo"), t.get("descKo")) for t in translated]
    except Exception as e:
        print(f"  번역 실패, 원문으로 폴백: {e}")
        return fallback


HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial")
THIN_BORDER = Border(*(Side(style="thin"),) * 4)
COLUMNS = ["번호", "종목명", "티커", "제목(원문)", "제목(한글)", "매체", "발행일", "URL", "내용요약(한글)"]


def write_xlsx(rows: list, path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "종목뉴스"

    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    for row in rows:
        ws.append(row)

    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if cell.column_letter == "I":  # 내용요약(한글)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    for col_idx, header in enumerate(COLUMNS, start=1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        width = max(15, min(60, len(header) * 3 + 10))
        ws.column_dimensions[letter].width = width

    wb.save(path)


def main():
    os.makedirs(NEWS_DIR, exist_ok=True)

    per_stock = {}
    for stock in FEATURED_STOCKS:
        print(f"[{stock['name']}] 뉴스 수집 중...")
        items = collect_for_stock(stock)
        per_stock[stock["code"]] = items
        print(f"  {len(items)}건 수집")

    all_items = [it for items in per_stock.values() for it in items]
    if not all_items:
        print("수집된 뉴스가 없어 종료")
        return

    translations = translate_batch(all_items)
    trans_map = {id(it): t for it, t in zip(all_items, translations)}

    rows = []
    row_no = 1
    for stock in FEATURED_STOCKS:
        for it in per_stock[stock["code"]]:
            title_ko, desc_ko = trans_map.get(id(it), (None, None))
            rows.append([
                row_no, stock["name"], stock["code"], it["title"],
                title_ko or it["title"], it["media"], it["date"], it["link"],
                desc_ko or it["desc"],
            ])
            row_no += 1

    run_date = datetime.now(KST).strftime("%Y%m%d")
    out_path = os.path.join(NEWS_DIR, f"종목뉴스_{run_date}.xlsx")
    write_xlsx(rows, out_path)
    print(f"\n완료: {len(rows)}건 저장 -> {out_path}")


if __name__ == "__main__":
    main()
